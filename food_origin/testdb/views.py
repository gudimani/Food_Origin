from django.shortcuts import render, redirect
from .models import Egg
from .tables import EggTable
from django.db.models import Q
import openpyxl
from django.http import HttpResponse
from django.core import serializers
import requests
from bs4 import BeautifulSoup
from django.urls import reverse
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages

def egg_table(request):
    queryset = Egg.objects.all()
    

    # Filtering
    filter_param = request.GET.get('filter_param')
    if filter_param:
        queryset = queryset.filter(Q(eggtype__icontains=filter_param) | Q(color__icontains=filter_param)| Q(size__icontains=filter_param)| Q(weight__icontains=filter_param))
        
    # # Searching
    # search_param = request.GET.get('search_param')
    # if search_param:
    #     queryset = queryset.filter(Q(eggtype__icontains=search_param) | Q(color__icontains=search_param)| Q(size__icontains=search_param)| Q(weight__icontains=search_param))

    # Sorting
    sort_param = request.GET.get('sort_param')
    if sort_param:
        queryset = queryset.order_by(sort_param)
    
    data = list(queryset.values())

    context = {
        'eggs': queryset
    }
    request.session['context'] = data
    return render(request, 'testdb/egg_table.html', context)


def export_table(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="egg_data.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set the column headers
    worksheet['A1'] = 'eggtype'
    worksheet['B1'] = 'color'
    worksheet['C1'] = 'size'
    worksheet['D1'] = 'weight'

    queryset_values = request.session.get('context')
    queryset = Egg.objects.filter(pk__in=[item['id'] for item in queryset_values])

    # Write data to the worksheet
    row_number = 2
    for obj in queryset:
        worksheet.cell(row=row_number, column=1, value=obj.eggtype)
        worksheet.cell(row=row_number, column=2, value=obj.color)
        worksheet.cell(row=row_number, column=3, value=obj.size)
        worksheet.cell(row=row_number, column=4, value=obj.weight)
        row_number += 1

    workbook.save(response)
    return response


def import_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        # Get the uploaded file from the request
        excel_file = request.FILES['excel_file']
        
        response = {
            'message':'success',
            'status':'200'
        }

        # Load the Excel file using openpyxl
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        row_count = 0
        error_list = []

        # Skip the header row if necessary
        # next(rows)
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):

            row_count += 1
            if len(row) < 4:
                error_list.append(f"All fields must be provided. Error at {row_count}")
                break
                # error_list['row_count' + row_count] = row_count
                # render(request, 'testdb/error.html', {'message': 'All fields must be provided.','row_count':row_count})

            if row[0] is None or row[1] is None or row[2] is None or row[3] is None:
                error_list.append(f"Some values are missing at {row_count}.")
                # error_list['row_count' + row_count] = row_count
                # return render(request, 'testdb/error.html', {'message': 'Some values are missing'})
            
            eggtype = row[0]
            color = row[1]
            size = row[2]
            weight = row[3]

            if Egg.objects.filter(eggtype=eggtype, color=color, size=size, weight=weight).exists():
                error_list.append(f"Duplicate row found in the database. Error at {row_count}.")
                # error_list['row_count' + row_count] = row_count
                # return render(request, 'testdb/error.html', {'message': 'Duplicate row found in the database.'})
            
            
            
            
        if len(error_list) >= 1:
            return render(request, 'testdb/error.html', {'errors': error_list})
           
        else:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                eggtype, color, size, weight = row
                Egg.objects.create(eggtype=eggtype, color=color, size=size, weight=weight)
                print("inserted")    
            
                  
        queryset = Egg.objects.all()

        context = {
        'eggs': queryset,
        'success':True
        }

    return render(request, 'testdb/egg_table.html', context)
    



def loginPage(request):
    # page = 'login'
    if request.user.is_authenticated:
        return redirect('egg_table')
    if request.method == 'POST':
        username = request.POST.get('username').lower()
        password = request.POST.get('password')
        try:
            user = User.objects.get(username=username)
        except:
            messages.error(request, 'User does not exist')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('egg_table')
        else:
            messages.error(request, 'Username or Password does not exist')
    # context = {'page': page}
    return redirect('home')


def home(request):
    return render(request, 'testdb/login_register.html')



